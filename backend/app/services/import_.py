import csv
import io
import uuid
from dataclasses import dataclass
from decimal import Decimal
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.errors import AppException
from app.models.grade import Grade
from app.services.grade import create_grade, update_grade
from app.services.user import create_student_account
from app.services.student import create_student as create_student_direct
from app.models.student import Student
from app.models.subject import Subject


@dataclass
class ImportResult:
    created: int
    skipped: int
    updated: int
    errors: list[dict[str, Any]]


async def import_students_csv(
    db: AsyncSession,
    *,
    teacher_id: uuid.UUID,
    school_id: uuid.UUID,
    class_id: uuid.UUID | None,
    content: bytes,
) -> ImportResult:
    created = 0
    skipped = 0
    updated = 0
    errors: list[dict[str, Any]] = []
    f = io.StringIO(content.decode("utf-8"))
    reader = csv.DictReader(f)
    required = {"email", "name", "student_number"}
    if not required.issubset(set(reader.fieldnames or [])):
        missing = list(required - set(reader.fieldnames or []))
        errors.append({"row": 0, "error": f"Missing columns: {', '.join(missing)}"})
        return ImportResult(created, skipped, updated, errors)

    from app.schemas.user import StudentCreate

    for idx, row in enumerate(reader, start=2):
        try:
            resolved_class_id = row.get("class_id") or (str(class_id) if class_id else None)
            if not resolved_class_id:
                errors.append({"row": idx, "error": "class_id query parameter is required"})
                skipped += 1
                continue
            data = StudentCreate(
                email=row["email"],
                name=row["name"],
                class_id=resolved_class_id,
                student_number=int(row["student_number"]),
                birth_date=row.get("birth_date") or None,
            )
            try:
                await create_student_account(db, school_id=school_id, teacher_id=teacher_id, data=data)
                created += 1
            except AppException as e:
                if e.code in {"USER_DUPLICATE_EMAIL", "STUDENT_DUPLICATE_NUMBER"}:
                    skipped += 1
                else:
                    errors.append({"row": idx, "error": f"{e.code}:{e.detail}"})
        except Exception as e:  # validation error
            errors.append({"row": idx, "error": str(e)})

    return ImportResult(created, skipped, updated, errors)


async def import_grades_csv(
    db: AsyncSession,
    *,
    teacher_id: uuid.UUID,
    class_id: uuid.UUID | None,
    semester_id: uuid.UUID | None,
    content: bytes,
) -> ImportResult:
    created = 0
    skipped = 0
    updated = 0
    errors: list[dict[str, Any]] = []
    f = io.StringIO(content.decode("utf-8"))
    reader = csv.DictReader(f)
    fieldnames = set(reader.fieldnames or [])
    legacy_required = {"student_id", "subject_id", "semester_id", "score"}
    prd_required = {"student_number", "subject_name", "score"}
    use_prd_contract = prd_required.issubset(fieldnames)
    if not use_prd_contract and not legacy_required.issubset(fieldnames):
        missing = list(prd_required - fieldnames)
        errors.append({"row": 0, "error": f"Missing columns: {', '.join(missing)}"})
        return ImportResult(created, skipped, updated, errors)

    students_by_number: dict[int, uuid.UUID] = {}
    subjects_by_name: dict[str, uuid.UUID] = {}
    if use_prd_contract:
        if class_id is None or semester_id is None:
            errors.append({"row": 0, "error": "class_id and semester_id query parameters are required"})
            return ImportResult(created, skipped, updated, errors)
        result = await db.execute(select(Student).where(Student.class_id == class_id))
        students_by_number = {student.student_number: student.id for student in result.scalars().all()}
        result = await db.execute(select(Subject).where(Subject.class_id == class_id))
        subjects_by_name = {subject.name: subject.id for subject in result.scalars().all()}

    for idx, row in enumerate(reader, start=2):
        try:
            try:
                if use_prd_contract:
                    resolved_student_id = students_by_number.get(int(row["student_number"]))
                    if resolved_student_id is None:
                        errors.append({"row": idx, "error": f"번호 {row['student_number']} 학생을 찾을 수 없습니다."})
                        skipped += 1
                        continue
                    resolved_subject_id = subjects_by_name.get(row["subject_name"])
                    if resolved_subject_id is None:
                        errors.append({"row": idx, "error": f"과목 '{row['subject_name']}'을(를) 찾을 수 없습니다."})
                        skipped += 1
                        continue
                    resolved_semester_id = semester_id
                else:
                    resolved_student_id = uuid.UUID(row["student_id"])
                    resolved_subject_id = uuid.UUID(row["subject_id"])
                    resolved_semester_id = uuid.UUID(row["semester_id"])

                await create_grade(
                    db,
                    student_id=resolved_student_id,
                    subject_id=resolved_subject_id,
                    semester_id=resolved_semester_id,
                    score=Decimal(row["score"]),
                    created_by=teacher_id,
                    teacher_id=teacher_id,
                )
                created += 1
            except AppException as e:
                if e.code == "GRADE_DUPLICATE":
                    grade_result = await db.execute(
                        select(Grade)
                        .where(Grade.student_id == resolved_student_id)
                        .where(Grade.subject_id == resolved_subject_id)
                        .where(Grade.semester_id == resolved_semester_id)
                    )
                    grade = grade_result.scalar_one_or_none()
                    if grade is None:
                        errors.append({"row": idx, "error": f"{e.code}:{e.detail}"})
                    else:
                        await update_grade(
                            db,
                            grade_id=grade.id,
                            score=Decimal(row["score"]),
                            teacher_id=teacher_id,
                        )
                        updated += 1
                else:
                    errors.append({"row": idx, "error": f"{e.code}:{e.detail}"})
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})

    return ImportResult(created, skipped, updated, errors)


# --- XLSX Import Functions ---


async def import_students_xlsx(
    db: AsyncSession,
    *,
    teacher_id: uuid.UUID,
    school_id: uuid.UUID,
    class_id: uuid.UUID,
    content: bytes,
) -> ImportResult:
    from openpyxl import load_workbook

    created = 0
    skipped = 0
    errors: list[dict[str, Any]] = []
    try:
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception as e:
        return ImportResult(0, 0, 0, [{"row": 0, "error": f"Invalid XLSX: {e}"}])

    all_rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))
    if not all_rows:
        return ImportResult(0, 0, 0, [{"row": 0, "error": "Empty worksheet"}])

    header = [str(c).strip() if c is not None else "" for c in all_rows[0]]
    expected = ["이름", "이메일", "번호", "생년월일", "성별", "연락처", "주소"]
    if header[: len(expected)] != expected:
        errors.append({"row": 1, "error": "Header mismatch"})
        return ImportResult(created, skipped, 0, errors)

    from datetime import date, datetime

    for row_idx, row in enumerate(all_rows[1:], start=2):
        name = (row[0] if len(row) > 0 else None) or None
        email = (row[1] if len(row) > 1 else None) or None
        number = row[2] if len(row) > 2 else None
        birth = row[3] if len(row) > 3 else None
        gender = row[4] if len(row) > 4 else None
        phone = row[5] if len(row) > 5 else None
        address = row[6] if len(row) > 6 else None

        if name is None or email is None or number is None:
            errors.append({"row": row_idx, "error": "이름/이메일/번호가 비어있습니다."})
            continue
        try:
            # Normalize birth
            birth_date = None
            if birth is not None and str(birth).strip() != "":
                if isinstance(birth, datetime):
                    birth_date = birth.date()
                elif isinstance(birth, date):
                    birth_date = birth
                else:
                    try:
                        birth_date = date.fromisoformat(str(birth))
                    except Exception:
                        birth_date = None

            gender_val = None
            if gender is not None:
                g = str(gender).strip().lower()
                if g in ("male", "남", "남자"):
                    gender_val = "male"
                elif g in ("female", "여", "여자"):
                    gender_val = "female"

            await create_student_direct(
                db,
                class_id=class_id,
                teacher_id=teacher_id,
                school_id=school_id,
                email=str(email),
                name=str(name),
                student_number=int(number),
                birth_date=birth_date,
                gender=gender_val,
                phone=None if phone in (None, "") else str(phone),
                address=None if address in (None, "") else str(address),
            )
            created += 1
        except AppException as e:
            if e.code == "STUDENT_DUPLICATE_NUMBER":
                errors.append({"row": row_idx, "error": f"{e.code}: {e.detail}"})
            else:
                errors.append({"row": row_idx, "error": f"{e.code}: {e.detail}"})
        except Exception as e:
            errors.append({"row": row_idx, "error": str(e)})

    return ImportResult(created, skipped, 0, errors)


async def import_grades_xlsx(
    db: AsyncSession,
    *,
    teacher_id: uuid.UUID,
    class_id: uuid.UUID,
    semester_id: uuid.UUID,
    content: bytes,
) -> ImportResult:
    from openpyxl import load_workbook

    created = 0
    skipped = 0
    errors: list[dict[str, Any]] = []

    try:
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception as e:
        return ImportResult(0, 0, 0, [{"row": 0, "error": f"Invalid XLSX: {e}"}])

    all_rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))
    if len(all_rows) < 2:
        return ImportResult(0, 0, 0, [{"row": 0, "error": "Empty worksheet"}])

    header = [str(c).strip() if c is not None else "" for c in all_rows[0]]
    if not header or header[0] != "번호":
        errors.append({"row": 1, "error": "첫 번째 열은 '번호'여야 합니다."})
        return ImportResult(created, skipped, 0, errors)

    # Map subject names to ids
    subjects_by_name: dict[str, uuid.UUID] = {}
    result = await db.execute(select(Subject).where(Subject.class_id == class_id))
    for s in result.scalars().all():
        subjects_by_name[s.name] = s.id

    subject_ids: list[tuple[int, uuid.UUID]] = []
    for idx, sname in enumerate(header[1:], start=1):
        if not sname:
            continue
        sid = subjects_by_name.get(sname)
        if sid is None:
            errors.append({"row": 1, "error": f"과목 '{sname}'을(를) 찾을 수 없습니다."})
        else:
            subject_ids.append((idx, sid))

    if not subject_ids:
        return ImportResult(created, skipped, 0, errors)

    # Student number -> student id
    result = await db.execute(select(Student).where(Student.class_id == class_id))
    students_by_number = {s.student_number: s.id for s in result.scalars().all()}

    for row_idx, row in enumerate(all_rows[1:], start=2):
        student_number = row[0] if row else None
        if student_number is None:
            errors.append({"row": row_idx, "error": "학생 번호가 비어있습니다."})
            continue
        try:
            sid = students_by_number.get(int(student_number))
        except Exception:
            errors.append({"row": row_idx, "error": "학생 번호가 유효하지 않습니다."})
            continue
        if sid is None:
            errors.append({"row": row_idx, "error": f"번호 {student_number} 학생을 찾을 수 없습니다."})
            continue

        for col_idx, subj_id in subject_ids:
            score_val = row[col_idx] if len(row) > col_idx else None
            if score_val is None or str(score_val).strip() == "":
                continue
            try:
                await create_grade(
                    db,
                    student_id=sid,
                    subject_id=subj_id,
                    semester_id=semester_id,
                    score=Decimal(str(score_val)),
                    created_by=teacher_id,
                    teacher_id=teacher_id,
                )
                created += 1
            except AppException as e:
                errors.append({"row": row_idx, "error": f"{e.code}: {e.detail}"})
            except Exception as e:
                errors.append({"row": row_idx, "error": str(e)})

    return ImportResult(created, skipped, 0, errors)


def generate_student_template() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "학생 목록"
    ws.append(["이름", "이메일", "번호", "생년월일", "성별", "연락처", "주소"])
    ws.append(["김철수", "kim@example.com", 1, "2010-03-15", "male", "010-1234-5678", "서울시 강남구"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_grade_template(subject_names: list[str]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "성적 입력"
    ws.append(["번호", *subject_names])
    ws.append([1, 85, 90, 78])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
